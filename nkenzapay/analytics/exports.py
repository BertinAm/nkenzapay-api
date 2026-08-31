"""Data exports.

Seven datasets, CSV or Excel, filtered before they are built — which matters,
because the admin screen shows the row count on the button and the count has to
mean the same thing the file contains.

Personal data is exported deliberately and narrowly. Payment proofs are never
included; an export is a spreadsheet, not a copy of the evidence store.
"""
from __future__ import annotations

import csv
import io
import logging

from django.utils import timezone

logger = logging.getLogger(__name__)

DATASETS = ("transactions", "users", "payments", "fees", "analytics",
            "login_activity", "website_activity")


def run_export(job_id):
    """Build the file. Queued on Celery when a broker is configured."""
    from .models import ExportJob

    try:
        from .tasks import build_export

        build_export.delay(job_id)
        return
    except Exception as exc:  # noqa: BLE001 - no broker in development
        logger.info("Building export %s inline: %s", job_id, exc)

    job = ExportJob.objects.filter(pk=job_id).first()
    if job is not None:
        build(job)


def build(job):
    from nkenzapay.common.storage import storage

    from .models import ExportJob

    job.state = ExportJob.RUNNING
    job.save(update_fields=["state"])
    try:
        sheets = {name: list(rows_for(name, job)) for name in job.datasets}
        job.row_count = sum(max(0, len(rows) - 1) for rows in sheets.values())

        if job.fmt == "excel":
            data = to_xlsx(sheets)
            key = f"exports/{job.pk}.xlsx"
            content_type = ("application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet")
        else:
            data = to_csv(sheets)
            key = f"exports/{job.pk}.csv"
            content_type = "text/csv"

        storage().save_bytes(key, data, content_type)
        job.storage_key = key
        job.state = ExportJob.READY
    except Exception as exc:  # noqa: BLE001 - the failure belongs on the job row
        logger.exception("Export %s failed", job.pk)
        job.state = ExportJob.FAILED
        job.error = str(exc)[:2000]
    job.finished_at = timezone.now()
    job.save()
    return job


def rows_for(dataset, job):
    """Yield a header row followed by data rows."""
    builder = {
        "transactions": _transactions,
        "users": _users,
        "payments": _payments,
        "fees": _fees,
        "analytics": _analytics,
        "login_activity": _login_activity,
        "website_activity": _website_activity,
    }[dataset]
    return builder(job)


def _range_filter(queryset, job, field="created_at"):
    if job.date_from:
        queryset = queryset.filter(**{f"{field}__date__gte": job.date_from})
    if job.date_to:
        queryset = queryset.filter(**{f"{field}__date__lte": job.date_to})
    return queryset


def _transactions(job):
    from nkenzapay.transactions.models import Transaction

    yield ["Reference", "Created", "Customer", "Email", "Direction", "Status",
           "Send amount", "Send currency", "Rate", "Converted", "Fee percent",
           "Fee amount", "Receive amount", "Receive currency", "Method",
           "Verified at", "Payout sent at", "Completed at"]

    queryset = _range_filter(
        Transaction.objects.select_related("user__profile", "collect_method"), job
    )
    status_filter = job.filters.get("status")
    if status_filter:
        queryset = queryset.filter(status__in=status_filter)

    for t in queryset.iterator():
        yield [
            t.reference, t.created_at.isoformat(), t.user.display_name, t.user.email,
            t.direction, t.status, str(t.send_amount), t.send_currency_id,
            str(t.rate_used), str(t.converted_amount), str(t.fee_percent),
            str(t.fee_amount), str(t.receive_amount), t.receive_currency_id,
            t.collect_method.label,
            t.verified_at.isoformat() if t.verified_at else "",
            t.payout_sent_at.isoformat() if t.payout_sent_at else "",
            t.confirmed_at.isoformat() if t.confirmed_at else "",
        ]


def _users(job):
    from nkenzapay.accounts.models import User

    yield ["ID", "Email", "Legal name", "WhatsApp", "Country", "Joined",
           "Email verified", "Marketing opt-in", "Suspended", "Transfers"]

    queryset = _range_filter(
        User.objects.select_related("profile"), job, field="date_joined"
    )
    for u in queryset.iterator():
        profile = getattr(u, "profile", None)
        yield [
            u.pk, u.email,
            profile.legal_name if profile else "",
            profile.whatsapp_display if profile else "",
            profile.country_id if profile else "",
            u.date_joined.isoformat(),
            "yes" if u.email_verified_at else "no",
            "yes" if u.marketing_opt_in else "no",
            "yes" if u.is_suspended else "no",
            u.transactions.count(),
        ]


def _payments(job):
    from nkenzapay.transactions.models import Transaction

    yield ["Reference", "Method", "Side", "Country", "Amount", "Currency",
           "Status", "Created"]
    queryset = _range_filter(Transaction.objects.select_related("collect_method"), job)
    for t in queryset.iterator():
        yield [t.reference, t.collect_method.label, t.collect_method.side,
               t.collect_method.country_id, str(t.send_amount), t.send_currency_id,
               t.status, t.created_at.isoformat()]


def _fees(job):
    from nkenzapay.transactions.models import Transaction

    yield ["Reference", "Created", "Fee percent", "Fee amount", "Currency",
           "Send amount", "Corridor"]
    queryset = _range_filter(
        Transaction.objects.select_related("corridor__source", "corridor__target"), job
    )
    for t in queryset.iterator():
        yield [t.reference, t.created_at.isoformat(), str(t.fee_percent),
               str(t.fee_amount), t.receive_currency_id, str(t.send_amount),
               f"{t.corridor.source_id}-{t.corridor.target_id}"]


def _analytics(job):
    from django.db.models import Count

    from .models import PageView

    yield ["Path", "Views", "Unique sessions"]
    queryset = _range_filter(PageView.objects.all(), job, field="at")
    rows = (
        queryset.values("path")
        .annotate(views=Count("id"), sessions=Count("session_key", distinct=True))
        .order_by("-views")
    )
    for row in rows:
        yield [row["path"], row["views"], row["sessions"]]


def _login_activity(job):
    from nkenzapay.accounts.models import LoginActivity

    yield ["User", "Email", "At", "IP", "Device", "New device", "Succeeded"]
    queryset = _range_filter(LoginActivity.objects.select_related("user"), job, field="at")
    for row in queryset.iterator():
        yield [row.user_id, row.user.email, row.at.isoformat(), row.ip or "",
               row.device_label, "yes" if row.is_new_device else "no",
               "yes" if row.succeeded else "no"]


def _website_activity(job):
    from .models import PageView

    yield ["At", "Path", "Device", "Source", "Referrer", "Session", "User"]
    queryset = _range_filter(PageView.objects.all(), job, field="at")
    for row in queryset.iterator():
        yield [row.at.isoformat(), row.path, row.device, row.source,
               row.referrer, row.session_key, row.user_id or ""]


def to_csv(sheets: dict) -> bytes:
    """One file. Several datasets are stacked with a titled blank line between,
    which reads better in Excel than seven separate downloads."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for index, (name, rows) in enumerate(sheets.items()):
        if index:
            writer.writerow([])
        writer.writerow([f"# {name}"])
        writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(sheets: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    book = Workbook()
    book.remove(book.active)
    for name, rows in sheets.items():
        sheet = book.create_sheet(title=name[:31])
        for row in rows:
            sheet.append(row)
        if sheet.max_row:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
        for column in sheet.columns:
            width = max((len(str(c.value or "")) for c in column), default=10)
            sheet.column_dimensions[column[0].column_letter].width = min(40, width + 2)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
